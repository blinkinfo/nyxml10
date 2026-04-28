"""Telegram command and callback-query handlers."""

from __future__ import annotations

import asyncio
import csv
import io
import html as _html
import logging
from datetime import datetime, timezone
from typing import Any

import openpyxl
from telegram import Update
from telegram.error import BadRequest
from telegram.ext import (
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

import config as cfg
from bot.formatters import (
    format_demo_recent_trades,
    format_demo_stats,
    format_help,
    format_model_compare,
    format_model_status,
    format_pattern_stats,
    format_recent_signals,
    format_recent_trades,
    format_redeem_preview,
    format_redeem_results,
    format_redemption_history,
    format_retrain_blocked,
    format_retrain_complete,
    format_rolling_wr_analytics,
    format_rolling_wr_dashboard,
    format_rolling_wr_history,
    format_rolling_wr_import_instructions,
    format_rolling_wr_import_preview,
    format_rolling_wr_import_success,
    format_rolling_wr_settings,
    format_set_threshold,
    format_set_down_threshold,
    format_signal_stats,
    format_status,
    format_threshold_analytics,
    format_threshold_policy_dashboard,
    format_trade_stats,
)
from bot.keyboards import (
    back_to_menu,
    down_override_keyboard,
    main_menu,
    ml_menu,
    ml_volatility_gate_confirm_keyboard,
    pattern_keyboard,
    redeem_confirm_keyboard,
    redeem_done_keyboard,
    retrain_blocked_keyboard,
    rolling_wr_back_keyboard,
    rolling_wr_import_preview_keyboard,
    rolling_wr_menu,
    rolling_wr_settings_keyboard,
    settings_keyboard,
    signal_filter_row,
    threshold_analytics_keyboard,
    threshold_cancel_keyboard,
    threshold_menu,
    threshold_mode_keyboard,
    threshold_policy_choice_keyboard,
    trade_filter_row,
)
from bot.middleware import auth_check
from db import queries
from polymarket import account as pm_account

log = logging.getLogger(__name__)
MAX_ML_THRESHOLD = 0.95


def _parse_rolling_wr_percent(raw: str) -> float:
    value = float(raw.strip().replace('%', ''))
    if value < 0 or value > 100:
        raise ValueError('out of range')
    return round(value, 4)


def _parse_positive_int(raw: str) -> int:
    value = int(float(raw.strip()))
    if value <= 0:
        raise ValueError('non-positive')
    return value


def _coerce_excel_preview_rows(rows: list[dict[str, Any]], window_size: int) -> list[dict[str, Any]]:
    deduped: dict[int, dict[str, Any]] = {}
    for row in rows:
        deduped[int(row['slot_timestamp'])] = row
    ordered = [deduped[k] for k in sorted(deduped)]
    return ordered[-window_size:]


def _extract_rolling_wr_preview_from_workbook(file_bytes: bytes, filename: str, window_size: int, config: dict[str, Any]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        raise ValueError('Workbook is empty')
    headers = [str(v).strip() if v is not None else '' for v in values[0]]
    header_map = {h: i for i, h in enumerate(headers)}
    if 'slot_start' not in header_map or 'is_win' not in header_map or ('model_side' not in header_map and 'side' not in header_map):
        raise ValueError('Required columns: slot_start, is_win, and model_side or side')

    rows_found = 0
    rejected = 0
    eligible: list[dict[str, Any]] = []
    for raw in values[1:]:
        if all(v is None or str(v).strip() == '' for v in raw):
            continue
        rows_found += 1
        slot_start = raw[header_map['slot_start']]
        side_idx = header_map.get('model_side', header_map.get('side'))
        original_side = raw[side_idx]
        is_win_value = raw[header_map['is_win']]
        if slot_start is None or original_side is None or is_win_value is None:
            rejected += 1
            continue
        side = str(original_side).strip()
        if side not in {'Up', 'Down'}:
            rejected += 1
            continue
        normalized = str(is_win_value).strip().lower()
        if normalized in {'1', '1.0', 'true', 'yes', 'win'}:
            is_correct = 1
        elif normalized in {'0', '0.0', 'false', 'no', 'loss'}:
            is_correct = 0
        else:
            try:
                is_correct = 1 if int(float(normalized)) == 1 else 0
            except Exception:
                rejected += 1
                continue
        slot_str = str(slot_start).strip()
        try:
            parsed_dt = datetime.fromisoformat(slot_str.replace('Z', '+00:00'))
            slot_timestamp = int(parsed_dt.timestamp())
        except Exception:
            try:
                parsed_dt = datetime.strptime(slot_str, '%Y-%m-%d %H:%M:%S')
                slot_timestamp = int(parsed_dt.replace(tzinfo=timezone.utc).timestamp())
            except Exception:
                rejected += 1
                continue
        winner_side = side if is_correct else ('Down' if side == 'Up' else 'Up')
        eligible.append({
            'slot_timestamp': slot_timestamp,
            'slot_start': slot_str,
            'signal_slug': None,
            'original_side': side,
            'winner_side': winner_side,
            'is_correct': is_correct,
        })

    trimmed = _coerce_excel_preview_rows(eligible, window_size)
    sample_size = len(trimmed)
    wins = sum(1 for row in trimmed if row['is_correct'] == 1)
    losses = sample_size - wins
    win_rate = round((wins / sample_size) * 100, 2) if sample_size else None
    ready = sample_size >= config['min_samples']
    if not config['enabled']:
        policy = 'FOLLOW'
        reason = 'rolling WR disabled'
    elif not ready:
        policy = 'SKIP' if config['skip_when_unready'] else 'FOLLOW'
        reason = 'warming up'
    elif win_rate is not None and win_rate > config['invert_above']:
        policy = 'INVERT'
        reason = 'above invert threshold'
    elif win_rate is not None and win_rate < config['follow_below']:
        policy = 'FOLLOW'
        reason = 'below follow threshold'
    else:
        policy = 'SKIP'
        reason = 'neutral zone'
    status = {
        'enabled': config['enabled'],
        'policy': policy,
        'win_rate': win_rate,
        'sample_size': sample_size,
        'window_size': window_size,
        'min_samples': config['min_samples'],
        'ready': ready,
        'wins': wins,
        'losses': losses,
        'follow_below': config['follow_below'],
        'invert_above': config['invert_above'],
        'skip_when_unready': config['skip_when_unready'],
        'reason': reason,
    }
    return {
        'filename': filename,
        'rows_found': rows_found,
        'eligible_rows': len(eligible),
        'rejected_rows': rejected,
        'rows': trimmed,
        'status': status,
    }


async def _render_rolling_wr_hub(update: Update) -> None:
    status = await queries.evaluate_rolling_wr()
    text = format_rolling_wr_dashboard(status)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=rolling_wr_menu())
    else:
        await update.message.reply_text(text, reply_markup=rolling_wr_menu(), parse_mode='HTML')


async def _render_rolling_wr_settings(update: Update) -> None:
    config = await queries.get_rolling_wr_config()
    status = await queries.evaluate_rolling_wr()
    text = format_rolling_wr_settings(config, status)
    kb = rolling_wr_settings_keyboard(config['enabled'], config['skip_when_unready'])
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode='HTML')


async def _render_rolling_wr_analytics(update: Update) -> None:
    data = await queries.get_rolling_wr_analytics()
    text = format_rolling_wr_analytics(data)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=rolling_wr_back_keyboard())
    else:
        await update.message.reply_text(text, reply_markup=rolling_wr_back_keyboard(), parse_mode='HTML')


async def _render_rolling_wr_history(update: Update) -> None:
    data = await queries.get_rolling_wr_analytics()
    text = format_rolling_wr_history(data)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=rolling_wr_back_keyboard())
    else:
        await update.message.reply_text(text, reply_markup=rolling_wr_back_keyboard(), parse_mode='HTML')


@auth_check
async def cmd_rolling_wr(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await _render_rolling_wr_hub(update)


def _parse_ml_threshold(raw: str) -> float:
    threshold = float(raw)
    if threshold > MAX_ML_THRESHOLD:
        raise ValueError("out of range")
    return threshold


def _parse_blocked_ranges(raw: str) -> list[tuple[float, float]] | None:
    ranges: list[tuple[float, float]] = []
    if not raw or not raw.strip() or raw.strip().lower() == "none":
        return ranges
    for part in raw.split(","):
        part = part.strip()
        if "-" not in part:
            return None
        lo_str, _, hi_str = part.partition("-")
        try:
            lo = float(lo_str.strip())
            hi = float(hi_str.strip())
        except ValueError:
            return None
        if not (0.0 <= lo <= 1.0 and 0.0 <= hi <= 1.0):
            return None
        if lo > hi:
            lo, hi = hi, lo
        ranges.append((lo, hi))
    return ranges


def _parse_threshold_bucket(raw: str) -> str:
    bucket = queries.truncate_probability_bucket(raw)
    if bucket is None:
        raise ValueError("invalid bucket")
    return bucket


_start_time: datetime = datetime.now(timezone.utc)
_poly_client: Any = None


def set_poly_client(client: Any) -> None:
    global _poly_client
    _poly_client = client


def set_start_time() -> None:
    global _start_time
    _start_time = datetime.now(timezone.utc)


def _uptime() -> str:
    delta = datetime.now(timezone.utc) - _start_time
    total_seconds = int(delta.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes}m"
    return f"{minutes}m"


async def _safe_edit(query, text, reply_markup=None, parse_mode="HTML"):
    try:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except BadRequest as e:
        if "not modified" not in str(e).lower():
            raise


@auth_check
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (
        "\U0001f916 <b>Welcome to AutoPoly!</b>\n\n"
        "BTC Up/Down 5-min trading bot for Polymarket.\n"
        "Select an option below:"
    )
    await update.message.reply_text(text, reply_markup=main_menu(), parse_mode="HTML")


@auth_check
async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    connected = False
    balance = None
    positions = []
    if _poly_client:
        connected = await pm_account.get_connection_status(_poly_client)
        balance = await pm_account.get_balance(_poly_client)
        positions = await pm_account.get_open_positions(_poly_client)
    autotrade = await queries.is_autotrade_enabled()
    auto_redeem = await queries.is_auto_redeem_enabled()
    trade_amount = await queries.get_trade_amount()
    last_sig = await queries.get_last_signal()
    last_sig_str = None
    if last_sig:
        ss = last_sig["slot_start"].split(" ")[-1] if " " in last_sig["slot_start"] else last_sig["slot_start"]
        last_sig_str = f"{ss} UTC ({last_sig.get('model_side') or last_sig['side']})"
    demo_trade = await queries.is_demo_trade_enabled()
    demo_bankroll = await queries.get_demo_bankroll() if demo_trade else None
    trade_mode = await queries.get_trade_mode()
    trade_pct = await queries.get_trade_pct()
    text = format_status(
        connected=connected,
        balance=balance,
        autotrade=autotrade,
        trade_amount=trade_amount,
        open_positions=len(positions),
        uptime_str=_uptime(),
        last_signal=last_sig_str,
        auto_redeem=auto_redeem,
        demo_trade_enabled=demo_trade,
        demo_bankroll=demo_bankroll,
        trade_mode=trade_mode,
        trade_pct=trade_pct,
    )
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=back_to_menu())
    else:
        await update.message.reply_text(text, reply_markup=back_to_menu(), parse_mode="HTML")


async def _render_signals(update: Update, limit: int | None, active: str) -> None:
    stats = await queries.get_signal_stats(limit=limit)
    label = {"10": "Last 10", "50": "Last 50", "all": "All Time"}[active]
    text = format_signal_stats(stats, label)
    recent = await queries.get_recent_signals(10)
    text += format_recent_signals(recent)
    kb = signal_filter_row(active)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


@auth_check
async def cmd_signals(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await _render_signals(update, limit=None, active="all")


async def _render_trades(update: Update, limit: int | None, active: str) -> None:
    stats = await queries.get_trade_stats(limit=limit)
    label = {"10": "Last 10", "50": "Last 50", "all": "All Time"}[active]
    text = format_trade_stats(stats, label)
    recent = await queries.get_recent_trades(10)
    text += format_recent_trades(recent)
    kb = trade_filter_row(active)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


@auth_check
async def cmd_trades(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await _render_trades(update, limit=None, active="all")


@auth_check
async def cmd_settings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    autotrade = await queries.is_autotrade_enabled()
    auto_redeem = await queries.is_auto_redeem_enabled()
    trade_amount = await queries.get_trade_amount()
    trade_mode = await queries.get_trade_mode()
    trade_pct = await queries.get_trade_pct()
    demo_trade = await queries.is_demo_trade_enabled()
    demo_bankroll = await queries.get_demo_bankroll()
    ml_volatility_gate_enabled = await queries.get_ml_volatility_gate_enabled()
    at_text = "ON" if autotrade else "OFF"
    mode_summary = f"{trade_pct:.1f}%" if trade_mode == "pct" else f"${trade_amount:.2f}"
    dt_text = "ON" if demo_trade else "OFF"
    text = (
        f"\u2699\ufe0f <b>Settings</b>\n"
        f"AutoTrade: {at_text}  |  Mode: {mode_summary}  |  Demo: {dt_text}\n"
        f"Threshold routing is active per bucket. Legacy global invert now only acts as real-mode fallback for FOLLOW buckets.\n"
        f"ML volatility gate: {'ON' if ml_volatility_gate_enabled else 'OFF'}"
    )
    kb = settings_keyboard(autotrade, trade_amount, auto_redeem, demo_trade, demo_bankroll, trade_mode, trade_pct, False, ml_volatility_gate_enabled)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


@auth_check
async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = format_help()
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=back_to_menu())
    else:
        await update.message.reply_text(text, reply_markup=back_to_menu(), parse_mode="HTML")


@auth_check
async def cmd_redeem(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    from core.redeemer import scan_and_redeem
    wallet = cfg.POLYMARKET_FUNDER_ADDRESS
    if not wallet:
        text = "\u274c <b>Redeem Error</b>\n\nPOLYMARKET_FUNDER_ADDRESS is not configured."
        if update.callback_query:
            await update.callback_query.answer()
            await _safe_edit(update.callback_query, text, reply_markup=back_to_menu())
        else:
            await update.message.reply_text(text, parse_mode="HTML", reply_markup=back_to_menu())
        return
    scanning_text = "\U0001f50d <b>Scanning wallet for redeemable positions...</b>"
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, scanning_text)
        sent = None
    else:
        sent = await update.message.reply_text(scanning_text, parse_mode="HTML")
    try:
        results = await scan_and_redeem(wallet, dry_run=True)
    except Exception:
        log.exception("cmd_redeem: scan_and_redeem raised unexpectedly")
        error_text = "\u274c <b>Scan failed</b>\n\nCould not fetch positions. Please try again."
        if update.callback_query:
            await _safe_edit(update.callback_query, error_text, reply_markup=back_to_menu())
        else:
            await sent.edit_text(error_text, parse_mode="HTML", reply_markup=back_to_menu())
        return
    context.user_data["redeem_preview"] = results
    text = format_redeem_preview(results)
    kb = redeem_confirm_keyboard() if results else back_to_menu()
    if update.callback_query:
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await sent.edit_text(text, parse_mode="HTML", reply_markup=kb)


@auth_check
async def cmd_redemptions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    stats = await queries.get_redemption_stats()
    recent = await queries.get_recent_redemptions(10)
    text = format_redemption_history(stats, recent)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=back_to_menu())
    else:
        await update.message.reply_text(text, reply_markup=back_to_menu(), parse_mode="HTML")


@auth_check
async def cmd_download_csv(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("Preparing CSV...")
    rows = await queries.get_all_signals_for_export()
    fieldnames = [
        "id", "slot_start", "side", "model_side", "entry_price", "is_win", "pattern",
        "ml_p_up", "ml_p_down", "ml_probability_bucket", "ml_probability_used",
        "threshold_policy_real", "threshold_policy_demo",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)
    await query.message.reply_document(
        document=io.BytesIO(buf.getvalue().encode()),
        filename="signals.csv",
        caption="\U0001f4e5 All signals export (CSV)",
    )


async def _reply_with_excel_export(
    query,
    rows: list[dict[str, Any]],
    headers: list[str],
    worksheet_title: str,
    filename: str,
    caption: str,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = worksheet_title
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await query.message.reply_document(document=buf, filename=filename, caption=caption)


@auth_check
async def cmd_download_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("Preparing Excel...")
    headers = [
        "id", "slot_start", "slot_timestamp", "side", "model_side", "entry_price", "is_win", "pattern",
        "ml_p_up", "ml_p_down", "ml_probability_bucket", "ml_probability_used",
        "threshold_policy_real", "threshold_policy_demo",
        "rolling_wr_policy", "rolling_wr_wr", "rolling_wr_window_size", "rolling_wr_sample_size",
        "rolling_wr_follow_below", "rolling_wr_invert_above", "rolling_wr_ready",
    ]
    await _reply_with_excel_export(
        query=query,
        rows=await queries.get_all_signals_for_export(),
        headers=headers,
        worksheet_title="Signals",
        filename="signals.xlsx",
        caption="\U0001f4e5 All signals export (Excel)",
    )


@auth_check
async def cmd_download_trades_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("Preparing Excel...")
    headers = [
        "id", "signal_id", "slot_start", "slot_end", "side", "entry_price", "amount_usdc",
        "order_id", "fill_price", "status", "retry_count", "outcome", "is_win", "pnl",
        "resolved_at", "routing_mode", "routing_policy", "original_side", "routed_side",
        "policy_bucket", "policy_probability", "rolling_wr_policy", "rolling_wr_wr",
        "rolling_wr_window_size", "rolling_wr_sample_size", "rolling_wr_ready",
    ]
    await _reply_with_excel_export(
        query=query,
        rows=await queries.get_all_real_trades_for_export(),
        headers=headers,
        worksheet_title="Trades",
        filename="trades.xlsx",
        caption="\U0001f4e5 All trades export (Excel)",
    )


@auth_check
async def cmd_download_demo_trades_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("Preparing Excel...")
    headers = [
        "id", "signal_id", "slot_start", "slot_end", "side", "entry_price", "amount_usdc",
        "order_id", "fill_price", "status", "retry_count", "outcome", "is_win", "pnl",
        "resolved_at", "routing_mode", "routing_policy", "original_side", "routed_side",
        "policy_bucket", "policy_probability", "rolling_wr_policy", "rolling_wr_wr",
        "rolling_wr_window_size", "rolling_wr_sample_size", "rolling_wr_ready",
    ]
    await _reply_with_excel_export(
        query=query,
        rows=await queries.get_all_demo_trades_for_export(),
        headers=headers,
        worksheet_title="Demo Trades",
        filename="demo_trades.xlsx",
        caption="\U0001f4e5 All demo trades export (Excel)",
    )


@auth_check
async def cmd_patterns(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    rows = await queries.get_pattern_stats()
    text = format_pattern_stats(rows)
    kb = pattern_keyboard()
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


@auth_check
async def cmd_download_pattern_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer("Preparing Excel...")
    rows = await queries.get_pattern_stats_for_export()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patterns"
    ws.append(["Pattern", "Total Trades", "Wins", "Losses", "Win%", "W/L Ratio", "Deployed USDC", "Net PnL", "ROI%", "Last Seen"])
    for r in rows:
        ws.append([
            r["pattern"], r["total_trades"], r["wins"], r["losses"], r["win_pct"],
            r["wl_ratio"] if r["wl_ratio"] != float("inf") else "inf",
            r["total_deployed"], r["net_pnl"], r["roi_pct"], r["last_seen"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await query.message.reply_document(
        document=buf,
        filename="pattern_performance.xlsx",
        caption="\U0001f4e5 Per-pattern stats export (Excel)",
    )


async def _render_demo_stats(update: Update, limit: int | None = None, active: str = "all") -> None:
    from bot.keyboards import demo_filter_row
    stats = await queries.get_demo_trade_stats(limit=limit)
    bankroll = await queries.get_demo_bankroll()
    label = {"10": "Last 10", "50": "Last 50", "all": "All Time"}[active]
    text = format_demo_stats(stats, bankroll, label)
    recent = await queries.get_recent_demo_trades(10)
    text += format_demo_recent_trades(recent)
    kb = demo_filter_row(active)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


@auth_check
async def cmd_demo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await _render_demo_stats(update, limit=None, active="all")


@auth_check
async def cmd_thresholds(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (
        "\U0001f500 <b>Threshold Routing</b>\n"
        "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
        "\u2502  \U0001f500 Per-bucket signal routing\n"
        "\u2502  Real and demo configured separately\n"
        "\u2502  Unset buckets \u2192 \U0001f7e2 FOLLOW by default\n"
        "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
    )
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=threshold_menu())
    else:
        await update.message.reply_text(text, reply_markup=threshold_menu(), parse_mode="HTML")


_POLICIES_PAGE_SIZE = 10
_ANALYTICS_PAGE_SIZE = 5


async def _render_threshold_policies(update: Update, mode: str, page: int = 1) -> None:
    all_rows = await queries.list_threshold_policies(mode=mode)
    total = len(all_rows)
    total_pages = max(1, (total + _POLICIES_PAGE_SIZE - 1) // _POLICIES_PAGE_SIZE)
    page = max(1, min(page, total_pages))
    start = (page - 1) * _POLICIES_PAGE_SIZE
    page_rows = all_rows[start:start + _POLICIES_PAGE_SIZE]
    text = format_threshold_policy_dashboard(mode, page_rows, page=page, total_pages=total_pages)
    kb = threshold_mode_keyboard(mode, page=page, total_pages=total_pages)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


async def _render_threshold_analytics(update: Update, mode: str, page: int = 1) -> None:
    all_rows = await queries.get_threshold_stats(mode)
    total = len(all_rows)
    total_pages = max(1, (total + _ANALYTICS_PAGE_SIZE - 1) // _ANALYTICS_PAGE_SIZE)
    page = max(1, min(page, total_pages))
    start = (page - 1) * _ANALYTICS_PAGE_SIZE
    page_rows = all_rows[start:start + _ANALYTICS_PAGE_SIZE]
    text = format_threshold_analytics(mode, page_rows, page=page, total_pages=total_pages)
    kb = threshold_analytics_keyboard(mode, page=page, total_pages=total_pages)
    if update.callback_query:
        await update.callback_query.answer()
        await _safe_edit(update.callback_query, text, reply_markup=kb)
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")


@auth_check
async def cmd_threshold_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    mode = "real"
    if context.args and context.args[0].lower() in {"real", "demo"}:
        mode = context.args[0].lower()
    await _render_threshold_analytics(update, mode, page=1)


async def callback_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    data = query.data
    if data == "cmd_menu":
        await query.answer()
        await _safe_edit(query, "\U0001f916 <b>AutoPoly Menu</b>\n\nSelect an option:", reply_markup=main_menu())
    elif data == "cmd_status":
        await cmd_status(update, context)
    elif data == "cmd_signals":
        await _render_signals(update, limit=None, active="all")
    elif data == "cmd_trades":
        await _render_trades(update, limit=None, active="all")
    elif data == "cmd_settings":
        await cmd_settings(update, context)
    elif data == "cmd_help":
        await cmd_help(update, context)
    elif data == "cmd_redeem":
        await cmd_redeem(update, context)
    elif data == "cmd_redemptions":
        await cmd_redemptions(update, context)
    elif data == "signals_10":
        await _render_signals(update, limit=10, active="10")
    elif data == "signals_50":
        await _render_signals(update, limit=50, active="50")
    elif data == "signals_all":
        await _render_signals(update, limit=None, active="all")
    elif data == "trades_10":
        await _render_trades(update, limit=10, active="10")
    elif data == "trades_50":
        await _render_trades(update, limit=50, active="50")
    elif data == "trades_all":
        await _render_trades(update, limit=None, active="all")
    elif data == "toggle_autotrade":
        current = await queries.is_autotrade_enabled()
        await queries.set_setting("autotrade_enabled", "false" if current else "true")
        await query.answer(f"AutoTrade {'OFF' if current else 'ON'}")
        await cmd_settings(update, context)
    elif data == "toggle_auto_redeem":
        current = await queries.is_auto_redeem_enabled()
        await queries.set_setting("auto_redeem_enabled", "false" if current else "true")
        await query.answer(f"Auto-Redeem {'OFF' if current else 'ON'}")
        await cmd_settings(update, context)
    elif data == "toggle_trade_mode":
        current_mode = await queries.get_trade_mode()
        new_mode = "pct" if current_mode == "fixed" else "fixed"
        await queries.set_setting("trade_mode", new_mode)
        await query.answer(f"Trade mode switched to {new_mode.upper()}")
        await cmd_settings(update, context)
    elif data == "change_amount":
        await query.answer()
        trade_mode = await queries.get_trade_mode()
        trade_pct = await queries.get_trade_pct()
        trade_amount = await queries.get_trade_amount()
        if trade_mode == "pct":
            await _safe_edit(query, f"\U0001f522 <b>Set Trade Percentage</b>\n\nCurrent: <b>{trade_pct:.1f}%</b>\n\nType the percentage to use per trade.")
            context.user_data["awaiting_trade_pct"] = True
        else:
            await _safe_edit(query, f"\U0001f4b5 <b>Set Trade Amount</b>\n\nCurrent: <b>${trade_amount:.2f}</b>\n\nType the new amount in USDC:")
            context.user_data["awaiting_amount"] = True
    elif data == "download_csv":
        await cmd_download_csv(update, context)
    elif data == "download_xlsx":
        await cmd_download_excel(update, context)
    elif data == "download_trades_xlsx":
        await cmd_download_trades_excel(update, context)
    elif data == "download_demo_trades_xlsx":
        await cmd_download_demo_trades_excel(update, context)
    elif data == "redeem_confirm":
        await _handle_redeem_confirm(update, context)
    elif data == "redeem_cancel":
        context.user_data.pop("redeem_preview", None)
        await query.answer("Cancelled.")
        await _safe_edit(query, "\u274c Redemption cancelled.", reply_markup=back_to_menu())
    elif data == "toggle_demo_trade":
        current = await queries.is_demo_trade_enabled()
        await queries.set_setting("demo_trade_enabled", "false" if current else "true")
        await query.answer(f"Demo Trade {'OFF' if current else 'ON'}")
        await cmd_settings(update, context)
    elif data == "set_demo_bankroll":
        await query.answer()
        demo_bankroll = await queries.get_demo_bankroll()
        await _safe_edit(query, f"\U0001f4b0 <b>Set Demo Bankroll</b>\n\nCurrent balance: <b>${demo_bankroll:.2f}</b>\n\nType the new bankroll amount in USDC:")
        context.user_data["awaiting_demo_bankroll"] = True
    elif data == "reset_demo_bankroll":
        await queries.reset_demo_bankroll(1000.00)
        await query.answer("Demo bankroll reset to $1000.00")
        await cmd_settings(update, context)
    elif data == "toggle_ml_volatility_gate":
        gate_enabled = await queries.get_ml_volatility_gate_enabled()
        if gate_enabled:
            await query.answer()
            await _safe_edit(query, "\u26a0 <b>Disable ML volatility gate?</b>", reply_markup=ml_volatility_gate_confirm_keyboard())
        else:
            await queries.set_ml_volatility_gate_enabled(True)
            await query.answer("ML volatility gate ON")
            await cmd_settings(update, context)
    elif data == "confirm_disable_ml_volatility_gate":
        await queries.set_ml_volatility_gate_enabled(False)
        await query.answer("ML volatility gate OFF")
        await cmd_settings(update, context)
    elif data == "cancel_disable_ml_volatility_gate":
        await query.answer("ML volatility gate kept ON")
        await cmd_settings(update, context)
    elif data == "rolling_wr_settings":
        await _render_rolling_wr_settings(update)
    elif data == "rolling_wr_analytics":
        await _render_rolling_wr_analytics(update)
    elif data == "rolling_wr_history":
        await _render_rolling_wr_history(update)
    elif data == "rolling_wr_import":
        context.user_data["awaiting_rolling_wr_import"] = True
        context.user_data.pop("rolling_wr_import_preview", None)
        await query.answer()
        config = await queries.get_rolling_wr_config()
        await _safe_edit(query, format_rolling_wr_import_instructions(config), reply_markup=rolling_wr_back_keyboard())
    elif data == "rolling_wr_toggle":
        config = await queries.get_rolling_wr_config()
        await queries.set_rolling_wr_enabled(not config["enabled"])
        await query.answer(f"Rolling WR {'ON' if not config['enabled'] else 'OFF'}")
        await _render_rolling_wr_settings(update)
    elif data == "rolling_wr_toggle_skip_unready":
        config = await queries.get_rolling_wr_config()
        await queries.set_rolling_wr_skip_when_unready(not config["skip_when_unready"])
        await query.answer(f"Warm-up mode {'SKIP' if not config['skip_when_unready'] else 'FOLLOW'}")
        await _render_rolling_wr_settings(update)
    elif data == "rolling_wr_reset_defaults":
        await queries.set_rolling_wr_enabled(True)
        await queries.set_rolling_wr_window_size(320)
        await queries.set_rolling_wr_min_samples(320)
        await queries.set_rolling_wr_follow_below(49.0)
        await queries.set_rolling_wr_invert_above(51.0)
        await queries.set_rolling_wr_skip_when_unready(True)
        context.user_data.pop("awaiting_rolling_wr_window", None)
        context.user_data.pop("awaiting_rolling_wr_min_samples", None)
        context.user_data.pop("awaiting_rolling_wr_follow_below", None)
        context.user_data.pop("awaiting_rolling_wr_invert_above", None)
        await query.answer("Rolling WR defaults restored")
        await _render_rolling_wr_settings(update)
    elif data == "rolling_wr_set_window":
        context.user_data["awaiting_rolling_wr_window"] = True
        context.user_data.pop("awaiting_rolling_wr_min_samples", None)
        context.user_data.pop("awaiting_rolling_wr_follow_below", None)
        context.user_data.pop("awaiting_rolling_wr_invert_above", None)
        config = await queries.get_rolling_wr_config()
        await query.answer()
        await _safe_edit(query, f"<b>Set Rolling WR Window</b>\n\nCurrent: <b>{config['window_size']}</b>\n\nSend a whole number for the rolling window size.", reply_markup=rolling_wr_back_keyboard())
    elif data == "rolling_wr_set_min_samples":
        context.user_data["awaiting_rolling_wr_min_samples"] = True
        context.user_data.pop("awaiting_rolling_wr_window", None)
        context.user_data.pop("awaiting_rolling_wr_follow_below", None)
        context.user_data.pop("awaiting_rolling_wr_invert_above", None)
        config = await queries.get_rolling_wr_config()
        await query.answer()
        await _safe_edit(query, f"<b>Set Rolling WR Min Samples</b>\n\nCurrent: <b>{config['min_samples']}</b>\nWindow: <b>{config['window_size']}</b>\n\nSend a whole number up to the window size.", reply_markup=rolling_wr_back_keyboard())
    elif data == "rolling_wr_set_follow_below":
        context.user_data["awaiting_rolling_wr_follow_below"] = True
        context.user_data.pop("awaiting_rolling_wr_window", None)
        context.user_data.pop("awaiting_rolling_wr_min_samples", None)
        context.user_data.pop("awaiting_rolling_wr_invert_above", None)
        config = await queries.get_rolling_wr_config()
        await query.answer()
        await _safe_edit(query, f"<b>Set Follow-Below Threshold</b>\n\nCurrent: <b>{config['follow_below']:.2f}%</b>\n\nSend a percentage between 0 and 100.", reply_markup=rolling_wr_back_keyboard())
    elif data == "rolling_wr_set_invert_above":
        context.user_data["awaiting_rolling_wr_invert_above"] = True
        context.user_data.pop("awaiting_rolling_wr_window", None)
        context.user_data.pop("awaiting_rolling_wr_min_samples", None)
        context.user_data.pop("awaiting_rolling_wr_follow_below", None)
        config = await queries.get_rolling_wr_config()
        await query.answer()
        await _safe_edit(query, f"<b>Set Invert-Above Threshold</b>\n\nCurrent: <b>{config['invert_above']:.2f}%</b>\n\nSend a percentage between 0 and 100.", reply_markup=rolling_wr_back_keyboard())
    elif data == "rolling_wr_import_confirm":
        preview = context.user_data.get("rolling_wr_import_preview")
        if not preview:
            await query.answer("No import preview found", show_alert=True)
            await _render_rolling_wr_hub(update)
        else:
            result = await queries.replace_rolling_wr_import(
                filename=preview.get("filename"),
                rows=preview.get("rows") or [],
                window_size_hint=(preview.get("status") or {}).get("window_size") or len(preview.get("rows") or []),
                notes=f"Imported from Telegram upload: {preview.get('filename') or 'signals.xlsx'}",
            )
            context.user_data.pop("rolling_wr_import_preview", None)
            context.user_data["awaiting_rolling_wr_import"] = False
            await query.answer("Rolling WR import applied")
            await _safe_edit(query, format_rolling_wr_import_success(result), reply_markup=rolling_wr_back_keyboard())
    elif data == "rolling_wr_import_cancel":
        context.user_data.pop("rolling_wr_import_preview", None)
        context.user_data["awaiting_rolling_wr_import"] = False
        await query.answer("Import cancelled")
        await _render_rolling_wr_hub(update)
    elif data == "cmd_demo":
        await _render_demo_stats(update, active="all")
    elif data == "demo_10":
        await _render_demo_stats(update, limit=10, active="10")
    elif data == "demo_50":
        await _render_demo_stats(update, limit=50, active="50")
    elif data == "demo_all":
        await _render_demo_stats(update, limit=None, active="all")
    elif data == "cmd_patterns":
        await cmd_patterns(update, context)
    elif data == "download_pattern_xlsx":
        await cmd_download_pattern_excel(update, context)
    elif data == "cmd_ml":
        await query.answer()
        await _safe_edit(query, "\U0001f916 <b>ML Model</b>", reply_markup=ml_menu())
    elif data == "cmd_thresholds":
        await cmd_thresholds(update, context)
    elif data == "cmd_rolling_wr":
        await cmd_rolling_wr(update, context)
    elif data == "thresholds_real" or (data.startswith("thresholds_real_p") and data[len("thresholds_real_p"):].isdigit()):
        page = int(data[len("thresholds_real_p"):]) if "_p" in data else 1
        await _render_threshold_policies(update, "real", page=page)
    elif data == "thresholds_demo" or (data.startswith("thresholds_demo_p") and data[len("thresholds_demo_p"):].isdigit()):
        page = int(data[len("thresholds_demo_p"):]) if "_p" in data else 1
        await _render_threshold_policies(update, "demo", page=page)
    elif data == "threshold_stats_real" or (data.startswith("threshold_stats_real_p") and data[len("threshold_stats_real_p"):].isdigit()):
        page = int(data[len("threshold_stats_real_p"):]) if "_p" in data else 1
        await _render_threshold_analytics(update, "real", page=page)
    elif data == "threshold_stats_demo" or (data.startswith("threshold_stats_demo_p") and data[len("threshold_stats_demo_p"):].isdigit()):
        page = int(data[len("threshold_stats_demo_p"):]) if "_p" in data else 1
        await _render_threshold_analytics(update, "demo", page=page)
    elif data == "noop":
        await query.answer()
    elif data.startswith("threshold_set_"):
        mode = data.split("_")[-1]
        context.user_data["awaiting_threshold_bucket"] = mode
        mode_emoji = "\U0001f4ca" if mode == "real" else "\U0001f9ea"
        prompt = (
            f"\u270f\ufe0f <b>Set Bucket Policy \u2014 {mode.upper()}</b>\n"
            "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            f"\u2502  {mode_emoji} Enter the bucket value to configure\n"
            "\u2502\n"
            "\u2502  Example:  <code>0.53</code>  or  <code>0.57</code>\n"
            "\u2502  Range:    0.00 \u2013 1.00\n"
            "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            "Reply with the bucket number now."
        )
        await _safe_edit(query, prompt, reply_markup=threshold_cancel_keyboard(mode))
    elif data.startswith("threshold_clear_"):
        mode = data.split("_")[-1]
        context.user_data["awaiting_threshold_clear_bucket"] = mode
        mode_emoji = "\U0001f4ca" if mode == "real" else "\U0001f9ea"
        prompt = (
            f"\U0001f5d1 <b>Clear Bucket Policy \u2014 {mode.upper()}</b>\n"
            "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            f"\u2502  {mode_emoji} Enter the bucket value to clear\n"
            "\u2502\n"
            "\u2502  Example:  <code>0.53</code>  or  <code>0.57</code>\n"
            "\u2502  Range:    0.00 \u2013 1.00\n"
            "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            "Reply with the bucket number now."
        )
        await _safe_edit(query, prompt, reply_markup=threshold_cancel_keyboard(mode))
    elif data.startswith("threshold_policy_"):
        _, _, mode, bucket, policy = data.split("_", 4)
        await queries.set_threshold_policy(bucket, mode, policy)
        _policy_emoji = {"FOLLOW": "\U0001f7e2", "BLOCK": "\U0001f534", "INVERT": "\U0001f501"}
        p_emoji = _policy_emoji.get(policy, "\u2022")
        confirm = (
            f"\u2705 <b>Policy Set \u2014 {mode.upper()}</b>\n"
            "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            f"\u2502  \U0001faa3 Bucket:  <b>{bucket}</b>\n"
            f"\u2502  \U0001f4cc Policy:  {p_emoji} {policy}\n"
            "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
        )
        await query.answer(f"{mode} bucket {bucket} \u2192 {policy}")
        await _safe_edit(query, confirm, reply_markup=threshold_mode_keyboard(mode))
    elif data == "ml_status":
        await query.answer(); await cmd_model_status(update, context)
    elif data == "ml_compare":
        await query.answer(); await cmd_model_compare(update, context)
    elif data == "ml_promote":
        await query.answer(); await cmd_promote_model(update, context)
    elif data == "ml_retrain":
        await query.answer(); await cmd_retrain(update, context)
    elif data == "ml_set_threshold":
        await query.answer(); await _safe_edit(query, f"\u2699\ufe0f <b>Set ML Threshold</b>\n\nCurrent threshold: <b>{await queries.get_ml_threshold():.3f}</b>")
        context.user_data["awaiting_ml_threshold"] = True
    elif data == "ml_set_down_threshold":
        await query.answer(); await _safe_edit(query, f"\u2699\ufe0f <b>Set ML DOWN Threshold</b>\n\nCurrent DOWN threshold: <b>{(await queries.get_ml_down_threshold()) or 0.0:.3f}</b>")
        context.user_data["awaiting_ml_down_threshold"] = True
    else:
        await query.answer("Unknown action")


async def _handle_redeem_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    from core.redeemer import redeem_position
    query = update.callback_query
    await query.answer("Executing redemptions...")
    preview = context.user_data.pop("redeem_preview", None)
    if not preview:
        await _safe_edit(query, "\u274c <b>Nothing to redeem</b>", reply_markup=back_to_menu())
        return
    wallet = cfg.POLYMARKET_FUNDER_ADDRESS
    if not wallet:
        await _safe_edit(query, "\u274c POLYMARKET_FUNDER_ADDRESS not configured.", reply_markup=back_to_menu())
        return
    results: list[dict] = []
    for pos in preview:
        result = await redeem_position(pos["condition_id"], pos.get("collateral_token"))
        merged = {**pos, **result, "dry_run": False}
        results.append(merged)
        try:
            await queries.insert_redemption(
                condition_id=pos["condition_id"],
                outcome_index=pos["outcome_index"],
                size=pos["size"],
                title=pos.get("title"),
                tx_hash=result.get("tx_hash"),
                status="success" if result.get("success") else "failed",
                error=result.get("error"),
                gas_used=result.get("gas_used"),
                dry_run=False,
            )
        except Exception:
            log.exception("Failed to persist redemption record for condition=%s", pos.get("condition_id"))
    await _safe_edit(query, format_redeem_results(results), reply_markup=redeem_done_keyboard())


async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get("awaiting_trade_pct"):
        context.user_data["awaiting_trade_pct"] = False
        raw = update.message.text.strip().replace("%", "")
        try:
            pct = float(raw)
            if pct <= 0 or pct > 100:
                raise ValueError("out of range")
        except ValueError:
            await update.message.reply_text("\u274c Invalid percentage.", parse_mode="HTML")
            return
        pct = round(pct, 2)
        await queries.set_setting("trade_pct", str(pct))
        await update.message.reply_text(f"\u2705 Trade percentage set to <b>{pct:.2f}%</b>", parse_mode="HTML")
        return

    if context.user_data.get("awaiting_rolling_wr_window"):
        context.user_data["awaiting_rolling_wr_window"] = False
        try:
            window_size = _parse_positive_int(update.message.text.strip())
        except ValueError:
            await update.message.reply_text("❌ Invalid window size. Send a whole number greater than 0.", parse_mode="HTML")
            context.user_data["awaiting_rolling_wr_window"] = True
            return
        await queries.set_rolling_wr_window_size(window_size)
        config = await queries.get_rolling_wr_config()
        await update.message.reply_text(
            f"✅ Rolling WR window set to <b>{config['window_size']}</b>. Min samples is now <b>{config['min_samples']}</b>.",
            parse_mode="HTML",
            reply_markup=rolling_wr_settings_keyboard(config['enabled'], config['skip_when_unready']),
        )
        return

    if context.user_data.get("awaiting_rolling_wr_min_samples"):
        context.user_data["awaiting_rolling_wr_min_samples"] = False
        try:
            min_samples = _parse_positive_int(update.message.text.strip())
        except ValueError:
            await update.message.reply_text("❌ Invalid min samples. Send a whole number greater than 0.", parse_mode="HTML")
            context.user_data["awaiting_rolling_wr_min_samples"] = True
            return
        await queries.set_rolling_wr_min_samples(min_samples)
        config = await queries.get_rolling_wr_config()
        await update.message.reply_text(
            f"✅ Rolling WR min samples set to <b>{config['min_samples']}</b>.",
            parse_mode="HTML",
            reply_markup=rolling_wr_settings_keyboard(config['enabled'], config['skip_when_unready']),
        )
        return

    if context.user_data.get("awaiting_rolling_wr_follow_below"):
        context.user_data["awaiting_rolling_wr_follow_below"] = False
        try:
            value = _parse_rolling_wr_percent(update.message.text)
        except ValueError:
            await update.message.reply_text("❌ Invalid percentage. Send a value from 0 to 100.", parse_mode="HTML")
            context.user_data["awaiting_rolling_wr_follow_below"] = True
            return
        await queries.set_rolling_wr_follow_below(value)
        config = await queries.get_rolling_wr_config()
        await update.message.reply_text(
            f"✅ Follow-below threshold set to <b>{config['follow_below']:.2f}%</b>.",
            parse_mode="HTML",
            reply_markup=rolling_wr_settings_keyboard(config['enabled'], config['skip_when_unready']),
        )
        return

    if context.user_data.get("awaiting_rolling_wr_invert_above"):
        context.user_data["awaiting_rolling_wr_invert_above"] = False
        try:
            value = _parse_rolling_wr_percent(update.message.text)
        except ValueError:
            await update.message.reply_text("❌ Invalid percentage. Send a value from 0 to 100.", parse_mode="HTML")
            context.user_data["awaiting_rolling_wr_invert_above"] = True
            return
        await queries.set_rolling_wr_invert_above(value)
        config = await queries.get_rolling_wr_config()
        await update.message.reply_text(
            f"✅ Invert-above threshold set to <b>{config['invert_above']:.2f}%</b>.",
            parse_mode="HTML",
            reply_markup=rolling_wr_settings_keyboard(config['enabled'], config['skip_when_unready']),
        )
        return

    if context.user_data.get("awaiting_demo_bankroll"):
        context.user_data["awaiting_demo_bankroll"] = False
        raw = update.message.text.strip().replace("$", "")
        try:
            amount = float(raw)
            if amount < 0:
                raise ValueError("negative")
        except ValueError:
            await update.message.reply_text("\u274c Invalid amount.")
            return
        amount = round(amount, 2)
        await queries.set_demo_bankroll(amount)
        await update.message.reply_text(f"\u2705 Demo bankroll set to <b>${amount:.2f}</b>", parse_mode="HTML")
        return

    if context.user_data.get("awaiting_ml_threshold"):
        context.user_data["awaiting_ml_threshold"] = False
        try:
            threshold = _parse_ml_threshold(update.message.text.strip())
        except ValueError:
            await update.message.reply_text(f"\u274c Invalid value. Enter a number up to {MAX_ML_THRESHOLD:.2f}.", parse_mode="HTML")
            return
        await queries.set_ml_threshold(threshold)
        await update.message.reply_text(format_set_threshold(threshold), parse_mode="HTML", reply_markup=ml_menu())
        return

    if context.user_data.get("awaiting_ml_down_threshold"):
        context.user_data["awaiting_ml_down_threshold"] = False
        try:
            threshold = _parse_ml_threshold(update.message.text.strip())
        except ValueError:
            await update.message.reply_text(f"\u274c Invalid value. Enter a number up to {MAX_ML_THRESHOLD:.2f}.", parse_mode="HTML")
            return
        await queries.set_ml_down_threshold(threshold)
        await update.message.reply_text(format_set_down_threshold(threshold), parse_mode="HTML", reply_markup=ml_menu())
        return

    if context.user_data.get("awaiting_threshold_bucket"):
        mode = context.user_data.pop("awaiting_threshold_bucket")
        try:
            bucket = _parse_threshold_bucket(update.message.text.strip())
        except ValueError:
            await update.message.reply_text(
                "\u274c <b>Invalid bucket.</b> Send a number between <code>0.00</code> and <code>1.00</code>.",
                parse_mode="HTML",
                reply_markup=threshold_cancel_keyboard(mode),
            )
            context.user_data["awaiting_threshold_bucket"] = mode
            return
        mode_emoji = "\U0001f4ca" if mode == "real" else "\U0001f9ea"
        policy_card = (
            f"\U0001f3af <b>Set Policy \u2014 {mode.upper()}</b>  \u2502  Bucket <b>{bucket}</b>\n"
            "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            "\u2502  Choose how signals in this bucket\n"
            "\u2502  should be routed:\n"
            "\u251c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            "\u2502  \U0001f7e2 FOLLOW  \u2014 trade as model says\n"
            "\u2502  \U0001f534 BLOCK   \u2014 skip this signal\n"
            "\u2502  \U0001f501 INVERT  \u2014 flip direction\n"
            "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
        )
        await update.message.reply_text(
            policy_card,
            parse_mode="HTML",
            reply_markup=threshold_policy_choice_keyboard(mode, bucket),
        )
        return

    if context.user_data.get("awaiting_threshold_clear_bucket"):
        mode = context.user_data.pop("awaiting_threshold_clear_bucket")
        try:
            bucket = _parse_threshold_bucket(update.message.text.strip())
        except ValueError:
            await update.message.reply_text(
                "\u274c <b>Invalid bucket.</b> Send a number between <code>0.00</code> and <code>1.00</code>.",
                parse_mode="HTML",
                reply_markup=threshold_cancel_keyboard(mode),
            )
            context.user_data["awaiting_threshold_clear_bucket"] = mode
            return
        await queries.clear_threshold_policy(bucket, mode)
        clear_confirm = (
            f"\U0001f5d1 <b>Policy Cleared \u2014 {mode.upper()}</b>\n"
            "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
            f"\u2502  \U0001faa3 Bucket:  <b>{bucket}</b>\n"
            f"\u2502  \u21a9\ufe0f Now defaults to \U0001f7e2 FOLLOW\n"
            "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
        )
        await update.message.reply_text(
            clear_confirm,
            parse_mode="HTML",
            reply_markup=threshold_mode_keyboard(mode),
        )
        return

    if not context.user_data.get("awaiting_amount"):
        return
    context.user_data["awaiting_amount"] = False
    raw = update.message.text.strip().replace("$", "")
    try:
        amount = float(raw)
        if amount <= 0:
            raise ValueError("non-positive")
    except ValueError:
        await update.message.reply_text("\u274c Invalid amount.")
        return
    amount = round(amount, 2)
    await queries.set_setting("trade_amount_usdc", str(amount))
    await update.message.reply_text(f"\u2705 Trade amount updated to <b>${amount:.2f}</b>", parse_mode="HTML")


async def document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.user_data.get("awaiting_rolling_wr_import"):
        return
    doc = update.message.document
    if not doc or not (doc.file_name or '').lower().endswith('.xlsx'):
        await update.message.reply_text("❌ Please upload an .xlsx export file.", parse_mode="HTML")
        return
    tg_file = await doc.get_file()
    file_bytes = await tg_file.download_as_bytearray()
    config = await queries.get_rolling_wr_config()
    try:
        preview = _extract_rolling_wr_preview_from_workbook(bytes(file_bytes), doc.file_name or 'signals.xlsx', config['window_size'], config)
    except Exception as exc:
        await update.message.reply_text(f"❌ Import failed: {_html.escape(str(exc))}", parse_mode="HTML", reply_markup=rolling_wr_back_keyboard())
        return
    context.user_data["rolling_wr_import_preview"] = preview
    context.user_data["awaiting_rolling_wr_import"] = False
    await update.message.reply_text(
        format_rolling_wr_import_preview(preview),
        parse_mode="HTML",
        reply_markup=rolling_wr_import_preview_keyboard(),
    )


def register(application) -> None:
    application.add_handler(CommandHandler("start", cmd_start))
    application.add_handler(CommandHandler("status", cmd_status))
    application.add_handler(CommandHandler("signals", cmd_signals))
    application.add_handler(CommandHandler("trades", cmd_trades))
    application.add_handler(CommandHandler("settings", cmd_settings))
    application.add_handler(CommandHandler("help", cmd_help))
    application.add_handler(CommandHandler("redeem", cmd_redeem))
    application.add_handler(CommandHandler("redemptions", cmd_redemptions))
    application.add_handler(CommandHandler("demo", cmd_demo))
    application.add_handler(CommandHandler("patterns", cmd_patterns))
    application.add_handler(CommandHandler("rolling_wr", cmd_rolling_wr))
    application.add_handler(CommandHandler("thresholds", cmd_thresholds))
    application.add_handler(CommandHandler("threshold_stats", cmd_threshold_stats))
    application.add_handler(CommandHandler("set_threshold", cmd_set_threshold))
    application.add_handler(CommandHandler("set_down_threshold", cmd_set_down_threshold))
    application.add_handler(CommandHandler("set_blocked_ranges", cmd_set_blocked_ranges))
    application.add_handler(CommandHandler("show_blocked_ranges", cmd_show_blocked_ranges))
    application.add_handler(CommandHandler("model_status", cmd_model_status))
    application.add_handler(CommandHandler("model_compare", cmd_model_compare))
    application.add_handler(CommandHandler("promote_model", cmd_promote_model))
    application.add_handler(CommandHandler("retrain", cmd_retrain))
    application.add_handler(CallbackQueryHandler(callback_router))
    application.add_handler(MessageHandler(filters.Document.ALL, document_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    async def _error_handler(update, context):
        import traceback
        err_text = "".join(traceback.format_exception(type(context.error), context.error, context.error.__traceback__))
        log.error("Unhandled Telegram error:\n%s", err_text)
        try:
            if cfg.TELEGRAM_CHAT_ID:
                short = err_text[-800:] if len(err_text) > 800 else err_text
                await context.bot.send_message(
                    chat_id=int(cfg.TELEGRAM_CHAT_ID),
                    text=f"&#x26A0;&#xFE0F; <b>Unhandled Bot Error</b>\n<pre>{_html.escape(short)}</pre>",
                    parse_mode="HTML",
                )
        except Exception:
            log.exception("Failed to send error notification to Telegram")

    application.add_error_handler(_error_handler)


@auth_check
async def cmd_set_threshold(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text(f"Usage: /set_threshold <value>\nMax: {MAX_ML_THRESHOLD:.2f}", parse_mode="HTML")
        return
    try:
        threshold = _parse_ml_threshold(context.args[0])
    except (ValueError, IndexError):
        await update.message.reply_text(f"Invalid value. Maximum allowed: {MAX_ML_THRESHOLD:.2f}", parse_mode="HTML")
        return
    await queries.set_ml_threshold(threshold)
    await update.message.reply_text(format_set_threshold(threshold), parse_mode="HTML")


@auth_check
async def cmd_set_down_threshold(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text(f"Usage: /set_down_threshold <value>\nMax: {MAX_ML_THRESHOLD:.2f}", parse_mode="HTML")
        return
    try:
        threshold = _parse_ml_threshold(context.args[0])
    except (ValueError, IndexError):
        await update.message.reply_text(f"Invalid value. Maximum allowed: {MAX_ML_THRESHOLD:.2f}", parse_mode="HTML")
        return
    await queries.set_ml_down_threshold(threshold)
    await update.message.reply_text(format_set_down_threshold(threshold), parse_mode="HTML")


@auth_check
async def cmd_set_blocked_ranges(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text(
            "Usage: /set_blocked_ranges <ranges>\nExample: <code>/set_blocked_ranges 0.20-0.22,0.40-0.42</code>",
            parse_mode="HTML",
        )
        return
    ranges_str = " ".join(context.args).strip()
    parsed = _parse_blocked_ranges(ranges_str)
    if parsed is None:
        await update.message.reply_text("Invalid format. Use comma-separated <code>low-high</code> pairs.", parse_mode="HTML")
        return
    await queries.set_blocked_threshold_ranges(parsed)
    await update.message.reply_text(
        "Legacy blocked threshold ranges updated. They now act only as fallback for FOLLOW buckets.",
        parse_mode="HTML",
        reply_markup=ml_menu(),
    )


@auth_check
async def cmd_show_blocked_ranges(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    ranges = await queries.get_blocked_threshold_ranges()
    if not ranges:
        text = "<b>Blocked Threshold Ranges</b>\nNone configured."
    else:
        lines = "\n".join(f"[{lo:.2f}, {hi:.2f}]" for lo, hi in ranges)
        text = f"<b>Blocked Threshold Ranges</b>\n{lines}\n\nUsed only as fallback for FOLLOW buckets."
    await update.message.reply_text(text, parse_mode="HTML", reply_markup=ml_menu())


@auth_check
async def cmd_model_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    from ml import model_store
    send = update.callback_query.message.reply_text if update.callback_query else update.message.reply_text
    meta = model_store.load_metadata("current")
    if meta is None:
        await send("No model trained yet. Use /retrain to train one.", parse_mode="HTML")
        return
    threshold = await queries.get_ml_threshold()
    await send(format_model_status("current", meta, threshold), parse_mode="HTML", reply_markup=back_to_menu())


@auth_check
async def cmd_model_compare(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    from ml import model_store
    send = update.callback_query.message.reply_text if update.callback_query else update.message.reply_text
    current_meta = model_store.load_metadata("current")
    candidate_meta = model_store.load_metadata("candidate")
    if current_meta is None:
        await send("No current model. Use /retrain to train one.", parse_mode="HTML")
        return
    if candidate_meta is None:
        await send("No candidate model. Use /retrain to generate a candidate.", parse_mode="HTML")
        return
    await send(format_model_compare(current_meta, candidate_meta), parse_mode="HTML", reply_markup=back_to_menu())


@auth_check
async def cmd_promote_model(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    from ml import model_store
    from core.strategies.ml_strategy import request_model_reload
    send = update.callback_query.message.reply_text if update.callback_query else update.message.reply_text
    if not model_store.has_model("candidate"):
        await send("No candidate model to promote. Use /retrain first.", parse_mode="HTML")
        return
    model_store.promote_candidate()
    try:
        await model_store.promote_candidate_in_db()
    except Exception:
        log.exception("cmd_promote_model: failed to persist promotion to DB")
    try:
        promoted = await model_store.load_model_from_db("current")
        if promoted:
            from core.strategies.ml_strategy import set_model
            set_model(promoted)
    except Exception:
        log.exception("cmd_promote_model: failed to preload promoted model into strategy")
    request_model_reload()
    meta = model_store.load_metadata("current")
    threshold = await queries.get_ml_threshold()
    await send(f"{format_model_status('current (promoted)', meta or {}, threshold)}\n\nCandidate promoted to current.", parse_mode="HTML")


@auth_check
async def cmd_retrain(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    send = update.callback_query.message.reply_text if update.callback_query else update.message.reply_text
    await send("Retraining started... estimated time ~5-8 min. I'll notify you when done.", parse_mode="HTML")
    asyncio.create_task(_retrain_background(context.application, cfg.TELEGRAM_CHAT_ID))


async def _retrain_background(application, chat_id) -> None:
    import asyncio as _asyncio
    import html as _html
    from pathlib import Path
    from ml import data_fetcher, features as feat_eng, trainer, model_store

    async def notify(text: str, reply_markup=None) -> None:
        try:
            await application.bot.send_message(chat_id=int(chat_id), text=text, parse_mode="HTML", reply_markup=reply_markup)
        except Exception as e:
            log.warning("_retrain_background: failed to send notification: %s", e)

    async def notify_document(path: str, caption: str) -> bool:
        try:
            with open(path, "rb") as fh:
                await application.bot.send_document(chat_id=int(chat_id), document=fh, filename=Path(path).name, caption=caption, parse_mode="HTML")
            return True
        except Exception:
            log.exception("_retrain_background: failed to send retrain report document")
            return False

    try:
        loop = _asyncio.get_event_loop()
        data = await _asyncio.wait_for(loop.run_in_executor(None, lambda: data_fetcher.fetch_all(months=9)), timeout=1500)
        df_feat = await _asyncio.wait_for(loop.run_in_executor(None, lambda: feat_eng.build_features(data["df5"], data["df15"], data["df1h"], data["funding"], data["cvd"])), timeout=1500)
        result = await _asyncio.wait_for(loop.run_in_executor(None, lambda: trainer.train(df_feat, slot="candidate")), timeout=1500)
        meta = model_store.load_metadata("candidate") or {}
        threshold = result.get("threshold", 0.535)
        down_threshold = result.get("down_threshold", round(1.0 - threshold, 4))
        report_info = result.get("report_info") or {}
        report_error = result.get("report_error")
        report_path = report_info.get("path")
        try:
            await queries.set_ml_threshold(threshold)
            await queries.set_ml_down_threshold(down_threshold)
        except Exception as thr_exc:
            log.warning("Retrain: failed to persist thresholds to DB: %s", thr_exc)
        try:
            await model_store.save_model_to_db(result["model"], "candidate", meta)
        except Exception as db_exc:
            log.warning("Retrain: failed to save candidate to DB: %s", db_exc)
        if result.get("blocked"):
            main_msg, risk_msg = format_retrain_blocked(meta, threshold)
            await notify(main_msg, reply_markup=retrain_blocked_keyboard())
            if risk_msg:
                await notify(risk_msg)
        else:
            main_msg, risk_msg = format_retrain_complete(meta, threshold)
            await notify(main_msg)
            if risk_msg:
                await notify(risk_msg)
            try:
                from ml import model_store as _ms
                from core.strategies.ml_strategy import request_model_reload as _req_reload, set_model as _set_model
                _ms.promote_candidate()
                try:
                    await _ms.promote_candidate_in_db()
                except Exception:
                    log.exception("_retrain_background: failed to persist auto-promotion to DB")
                try:
                    _promoted = await _ms.load_model_from_db("current")
                    if _promoted:
                        _set_model(_promoted)
                except Exception:
                    log.exception("_retrain_background: failed to preload auto-promoted model")
                _req_reload()
            except Exception:
                log.exception("_retrain_background: auto-promotion failed")
        if report_path and not report_error:
            sent = await notify_document(report_path, "Detailed val/test trade ledger and hourly UTC stats from the completed retrain.")
            if not sent:
                await notify("<i>Retrain finished, but the Excel report could not be sent to Telegram.</i>")
    except _asyncio.TimeoutError:
        await notify("Retrain timed out after 25 min. Try again or check Railway logs.")
    except Exception as exc:
        log.exception("Retrain background task failed: %s", exc)
        await notify(f"\u274c <b>Retrain failed</b>\n\n{_html.escape(str(exc))}\n\nCheck Railway logs for details.")
